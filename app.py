# -*- coding: utf-8 -*-
"""
나라장터 입찰정보 수집 웹앱 (Streamlit) — 검색조건 확장판
검색조건: 검색유형(입찰공고/개찰결과), 기간, 업무구분, 참가제한지역, 추정가격
+ 입찰공고는 공고서 첨부파일에서 실무담당 전화번호 자동 추출
"""

import re
import io
import os
import ssl
import sys
import time
import zlib
import zipfile
import struct
import smtplib
import datetime
import threading
import subprocess
import urllib.parse
from email.message import EmailMessage

import requests
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill

st.set_page_config(page_title="나라장터 입찰정보 수집기", page_icon="📋",
                   layout="centered")

# ------------------------------------------------------------
# 코드표
# ------------------------------------------------------------
# 참가제한지역 (나라장터 검색화면과 동일한 목록, 2026 행정구역 반영)
# 값: 참가제한지역 코드 목록. 행정구역 개편(통합·승격) 전 공고까지 잡히도록
#     구 코드를 함께 조회한다. None이면 지역 필터 없음.
REGIONS = {
    "전체": None,
    "전국(제한없음)": [],            # 코드 필터 없이 조회 후 지역명으로 판별
    "서울특별시": ["11"],
    "전남광주통합특별시": ["46", "29"],   # 통합 전 전남(46)·광주(29) 코드 포함
    "부산광역시": ["26"],
    "대구광역시": ["27"],
    "인천광역시": ["28"],
    "대전광역시": ["30"],
    "울산광역시": ["31"],
    "세종특별자치시": ["36"],
    "경기도": ["41"],
    "충청북도": ["43"],
    "충청남도": ["44"],
    "경상북도": ["47"],
    "경상남도": ["48"],
    "제주특별자치도": ["50"],
    "강원특별자치도": ["51", "42"],      # 특별자치도 승격 전 코드(42) 포함
    "전북특별자치도": ["52", "45"],      # 특별자치도 승격 전 코드(45) 포함
}

# 공고의 '참가가능지역명'과 매칭하기 위한 지역명 키워드
RGN_NAME_KEYWORDS = {
    "전국(제한없음)": ["전국", "제한없음"],
    "서울특별시": ["서울"],
    "전남광주통합특별시": ["전남", "전라남도", "광주"],
    "부산광역시": ["부산"],
    "대구광역시": ["대구"],
    "인천광역시": ["인천"],
    "대전광역시": ["대전"],
    "울산광역시": ["울산"],
    "세종특별자치시": ["세종"],
    "경기도": ["경기"],
    "충청북도": ["충북", "충청북도"],
    "충청남도": ["충남", "충청남도"],
    "경상북도": ["경북", "경상북도"],
    "경상남도": ["경남", "경상남도"],
    "제주특별자치도": ["제주"],
    "강원특별자치도": ["강원"],
    "전북특별자치도": ["전북", "전라북도"],
}

# 참가가능지역 정보를 조회할 수 없는 공고용 예비 필터:
# 수요기관명에 지역 소속 시·군 이름이 있으면 해당 지역으로 간주
INSTT_KEYWORDS = {
    "서울특별시": ["서울"],
    "부산광역시": ["부산"],
    "대구광역시": ["대구", "군위"],
    "인천광역시": ["인천", "강화", "옹진"],
    "대전광역시": ["대전"],
    "울산광역시": ["울산"],
    "세종특별자치시": ["세종"],
    "경기도": ["경기", "수원", "성남", "의정부", "안양", "부천", "광명", "평택",
             "동두천", "안산", "고양", "과천", "구리", "남양주", "오산", "시흥",
             "군포", "의왕", "하남", "용인", "파주", "이천", "안성", "김포",
             "화성", "광주시", "양주", "포천", "여주", "연천", "가평", "양평"],
    "강원특별자치도": ["강원", "춘천", "원주", "강릉", "동해", "태백", "속초",
             "삼척", "홍천", "횡성", "영월", "평창", "정선", "철원", "화천",
             "양구", "인제", "고성", "양양"],
    "충청북도": ["충북", "충청북도", "청주", "충주", "제천", "보은", "옥천",
             "영동", "증평", "진천", "괴산", "음성", "단양"],
    "충청남도": ["충남", "충청남도", "천안", "공주", "보령", "아산", "서산",
             "논산", "계룡", "당진", "금산", "부여", "서천", "청양", "홍성",
             "예산", "태안"],
    "전북특별자치도": ["전북", "전라북도", "전주", "군산", "익산", "정읍", "남원",
             "김제", "완주", "진안", "무주", "장수", "임실", "순창", "고창", "부안"],
    "전남광주통합특별시": ["전남", "전라남도", "광주", "목포", "여수", "순천",
             "나주", "광양", "담양", "곡성", "구례", "고흥", "보성", "화순",
             "장흥", "강진", "해남", "영암", "무안", "함평", "영광", "장성",
             "완도", "진도", "신안"],
    "경상북도": ["경북", "경상북도", "포항", "경주", "김천", "안동", "구미",
             "영주", "영천", "상주", "문경", "경산", "의성", "청송", "영양",
             "영덕", "청도", "고령", "성주", "칠곡", "예천", "봉화", "울진", "울릉"],
    "경상남도": ["경남", "경상남도", "창원", "진주", "통영", "사천", "김해",
             "밀양", "거제", "양산", "의령", "함안", "창녕", "고성", "남해",
             "하동", "산청", "함양", "거창", "합천"],
    "제주특별자치도": ["제주", "서귀포"],
}

# 업무구분 -> (입찰공고 오퍼레이션, 개찰결과 오퍼레이션)
TASKS = {
    "공사":  ("getBidPblancListInfoCnstwkPPSSrch", "getOpengResultListInfoCnstwk"),
    "물품":  ("getBidPblancListInfoThngPPSSrch",   "getOpengResultListInfoThng"),
    "용역":  ("getBidPblancListInfoServcPPSSrch",  "getOpengResultListInfoServc"),
    "외자":  ("getBidPblancListInfoFrgcptPPSSrch", "getOpengResultListInfoFrgcpt"),
}

# 업무구분 -> 기초금액(추정가격) 오퍼레이션
# ※ 공사 등은 공고 목록 응답에 추정가격이 없고 이 오퍼레이션에 분리되어 있음
TASKS_BSIS = {
    "공사": "getBidPblancListInfoCnstwkBsisAmount",
    "용역": "getBidPblancListInfoServcBsisAmount",
    "물품": "getBidPblancListInfoThngBsisAmount",
    "외자": None,
}

BID_BASES = [  # 입찰공고정보서비스 (신/구 주소)
    "http://apis.data.go.kr/1230000/ad/BidPublicInfoService",
    "http://apis.data.go.kr/1230000/BidPublicInfoService05",
    "http://apis.data.go.kr/1230000/BidPublicInfoService04",
    "http://apis.data.go.kr/1230000/BidPublicInfoService",
]
SCSBID_BASES = [  # 낙찰정보서비스 (신/구 주소)
    "http://apis.data.go.kr/1230000/as/ScsbidInfoService",
    "http://apis.data.go.kr/1230000/ScsbidInfoService01",
    "http://apis.data.go.kr/1230000/ScsbidInfoService",
]

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
PHONE_RE = re.compile(r"0\d{1,2}[\s\-\.\)]{0,2}\d{3,4}[\s\-\.]{0,2}\d{4}")
CONTACT_KEYWORDS = ["문의", "담당", "감독", "실무", "연락처", "주무관",
                    "문의처", "담당자", "감독관", "공사감독", "전화",
                    "☎", "TEL", "Tel", "T.", "℡"]


# ------------------------------------------------------------
# 공통 함수
# ------------------------------------------------------------
def get_service_key():
    try:
        k = st.secrets["SERVICE_KEY"].strip()
    except Exception:
        return None
    if "%" in k:
        k = urllib.parse.unquote(k)
    return k


def call_api(bases, operation, extra_params, log):
    """여러 base URL을 순서대로 시도하며 전체 페이지 수집"""
    key = get_service_key()
    last_error = ""
    for base in bases:
        url = f"{base}/{operation}"
        items, page = [], 1
        try:
            while True:
                params = {"serviceKey": key, "pageNo": page, "numOfRows": 100,
                          "type": "json", **extra_params}
                r = requests.get(url, params=params, headers=HEADERS, timeout=30)
                if r.status_code != 200:
                    last_error = f"HTTP {r.status_code} ({url})"
                    break
                if r.text.lstrip().startswith("<"):
                    last_error = r.text[:300]
                    break
                body = r.json().get("response", {}).get("body", {})
                rows = body.get("items", [])
                if isinstance(rows, dict):
                    rows = rows.get("item", [])
                if not rows:
                    break
                items.extend(rows)
                total = int(body.get("totalCount", 0))
                log.write(f"목록 수신 중... {len(items)}/{total}건")
                if len(items) >= total:
                    break
                page += 1
                time.sleep(0.3)
            if items:
                return items, None
        except Exception as e:
            last_error = str(e)
            continue
    return [], last_error


def fetch_notice(bid_ntce_no, task, cache={}):
    """공고번호로 원 입찰공고 1건 조회 (개찰결과 → 담당자 연락처 추출용)"""
    if not bid_ntce_no:
        return None
    if bid_ntce_no in cache:
        return cache[bid_ntce_no]
    op = TASKS[task][0]
    key = get_service_key()
    item = None
    for base in BID_BASES:
        try:
            r = requests.get(f"{base}/{op}",
                             params={"serviceKey": key, "pageNo": 1,
                                     "numOfRows": 10, "type": "json",
                                     "inqryDiv": 2, "bidNtceNo": bid_ntce_no},
                             headers=HEADERS, timeout=30)
            if r.status_code != 200 or r.text.lstrip().startswith("<"):
                continue
            body = r.json().get("response", {}).get("body", {})
            rows = body.get("items", [])
            if isinstance(rows, dict):
                rows = rows.get("item", [])
            if rows:
                item = rows[-1]  # 재공고 등 차수가 여러 개면 최신 차수
                break
        except Exception:
            continue
    cache[bid_ntce_no] = item
    return item


def call_api_chunked(bases, operation, base_params, d_from, d_to, log,
                     chunk_days=28):
    """조회기간 제한(약 1개월)이 있는 오퍼레이션을 기간을 쪼개 전량 수집"""
    items, seen, last_err = [], set(), None
    cur = d_from
    while cur <= d_to:
        end = min(cur + datetime.timedelta(days=chunk_days - 1), d_to)
        params = {**base_params,
                  "inqryBgnDt": cur.strftime("%Y%m%d") + "0000",
                  "inqryEndDt": end.strftime("%Y%m%d") + "2359"}
        part, err = call_api(bases, operation, params, log)
        last_err = err or last_err
        for it in part:
            k = (it.get("bidNtceNo"), it.get("bidNtceOrd"))
            if k not in seen:
                seen.add(k)
                items.append(it)
        cur = end + datetime.timedelta(days=1)
    return items, (None if items else last_err)


def fetch_bsis_price(bid_ntce_no, task, cache={}):
    """공고번호로 기초금액 오퍼레이션에서 추정가격/기초금액 조회 (없으면 0)"""
    op = TASKS_BSIS.get(task)
    if not op or not bid_ntce_no:
        return 0
    ck = (task, bid_ntce_no)
    if ck in cache:
        return cache[ck]
    key = get_service_key()
    price = 0
    for base in BID_BASES:
        try:
            r = requests.get(f"{base}/{op}",
                             params={"serviceKey": key, "pageNo": 1,
                                     "numOfRows": 10, "type": "json",
                                     "inqryDiv": 2, "bidNtceNo": bid_ntce_no},
                             headers=HEADERS, timeout=30)
            if r.status_code != 200 or r.text.lstrip().startswith("<"):
                continue
            resp = r.json().get("response", {})
            if str(resp.get("header", {}).get("resultCode", "")) not in ("00", "0"):
                continue
            rows = resp.get("body", {}).get("items", [])
            if isinstance(rows, dict):
                rows = rows.get("item", [])
            if isinstance(rows, dict):
                rows = [rows]
            for it in rows:
                p = price_of(it)
                if p:
                    price = p
                    break
            break
        except Exception:
            continue
    cache[ck] = price
    return price


def openg_date_of(it):
    """항목의 개찰일시를 date로 파싱 (실패 시 None)"""
    s = str(it.get("opengDt", "") or it.get("rlOpengDt", ""))[:10]
    s = s.replace(".", "-").replace("/", "-")
    try:
        return datetime.date.fromisoformat(s)
    except ValueError:
        return None


def fetch_openg(bid_ntce_no, task, cache={}):
    """공고번호로 개찰결과 1건 조회 (낙찰업체·낙찰률 등)"""
    if not bid_ntce_no:
        return None
    ck = (task, bid_ntce_no)
    if ck in cache:
        return cache[ck]
    op = TASKS[task][1]
    key = get_service_key()
    item = None
    for base in SCSBID_BASES:
        try:
            r = requests.get(f"{base}/{op}",
                             params={"serviceKey": key, "pageNo": 1,
                                     "numOfRows": 10, "type": "json",
                                     "inqryDiv": 2, "bidNtceNo": bid_ntce_no},
                             headers=HEADERS, timeout=30)
            if r.status_code != 200 or r.text.lstrip().startswith("<"):
                continue
            resp = r.json().get("response", {})
            if str(resp.get("header", {}).get("resultCode", "")) not in ("00", "0"):
                continue
            rows = resp.get("body", {}).get("items", [])
            if isinstance(rows, dict):
                rows = rows.get("item", [])
            if isinstance(rows, dict):
                rows = [rows]
            if rows:
                item = rows[-1]
                break
        except Exception:
            continue
    cache[ck] = item
    return item


def fetch_psbl_rgn(bid_ntce_no, cache={}):
    """공고번호로 참가가능(제한)지역명 목록 조회.
    반환: 지역명 리스트 / [] (정상응답이며 제한 미지정) / None (조회 불가)"""
    if not bid_ntce_no:
        return None
    if bid_ntce_no in cache:
        return cache[bid_ntce_no]
    key = get_service_key()
    result = None
    for base in BID_BASES:
        try:
            r = requests.get(f"{base}/getBidPblancListInfoPrtcptPsblRgn",
                             params={"serviceKey": key, "pageNo": 1,
                                     "numOfRows": 30, "type": "json",
                                     "inqryDiv": 2, "bidNtceNo": bid_ntce_no},
                             headers=HEADERS, timeout=30)
            if r.status_code != 200 or r.text.lstrip().startswith("<"):
                continue
            resp = r.json().get("response", {})
            code = str(resp.get("header", {}).get("resultCode", ""))
            if code not in ("00", "0"):     # 오류 응답은 '조회 불가'로 처리
                continue
            body = resp.get("body", {})
            rows = body.get("items", [])
            if isinstance(rows, dict):
                rows = rows.get("item", [])
            if isinstance(rows, dict):
                rows = [rows]
            names = []
            for it in rows:
                if not isinstance(it, dict):
                    continue
                for f in ("prtcptPsblRgnNm", "prtcptLmtRgnNm", "rgnNm",
                          "prtcptPsblRgnCdNm"):
                    v = str(it.get(f, "") or "").strip()
                    if v:
                        names.append(v)
            # 정상응답이지만 행이 0건이면 totalCount로 진위 확인
            total = str(body.get("totalCount", "0"))
            if not names and total not in ("0", ""):
                result = None       # 행은 있다는데 지역명을 못 읽음 → 판단 보류
            else:
                result = names      # []: 제한 미지정(전국) 확정
            break
        except Exception:
            continue
    cache[bid_ntce_no] = result
    return result


def region_match(region_name, rgn_names, item):
    """나라장터 참가제한지역 필터와 동일한 판정.
    rgn_names: 공고의 참가가능지역명 목록(None=조회실패, []=제한없음)"""
    if region_name == "전체":
        return True
    if rgn_names is None:
        # 지역정보 조회 실패 → 수요기관명으로 예비 판정
        if region_name == "전국(제한없음)":
            return False
        kws = INSTT_KEYWORDS.get(region_name, [region_name])
        blob = str(item.get("dminsttNm", "")) + str(item.get("ntceInsttNm", ""))
        return any(k in blob for k in kws)
    if not rgn_names:          # 제한지역 미지정 = 전국(제한없음)
        return region_name == "전국(제한없음)"
    joined = " ".join(rgn_names)
    if region_name == "전국(제한없음)":
        return any(k in joined for k in RGN_NAME_KEYWORDS["전국(제한없음)"])
    return any(k in joined for k in RGN_NAME_KEYWORDS.get(region_name, [region_name]))


def price_of(item):
    for f in ("presmptPrce", "asignBdgtAmt", "bdgtAmt", "sucsfbidAmt",
              "bssamt", "bssAmt", "bssAmount"):
        v = str(item.get(f, "") or "").replace(",", "")
        if v.replace(".", "").isdigit():
            return int(float(v))
    return 0


def attachments_of(item, limit):
    """첨부파일 중 전화번호가 있을 법한 문서만, 공고문 우선순으로 선별"""
    SKIP_EXT = (".dwg", ".dxf", ".zip", ".7z", ".egg", ".jpg", ".jpeg",
                ".png", ".gif", ".tif", ".bmp", ".xls", ".xlsx", ".xlsm")
    HIGH = ("공고문", "입찰공고", "공고서", "재공고")        # 담당자 정보 최다
    MID = ("유의서", "현장설명", "설명서", "안내", "과업")   # 그 다음
    LOW_NAME = ("도면", "내역", "산출", "수량", "단가", "규격서")  # 사실상 無
    scored = []
    for i in range(1, 11):
        u = item.get(f"ntceSpecDocUrl{i}")
        n = item.get(f"ntceSpecFileNm{i}") or f"file{i}"
        if not u:
            continue
        low = n.lower()
        if low.endswith(SKIP_EXT):
            continue                       # 파싱 불가/전화번호 없음 → 다운로드 생략
        s = 0
        if any(k in n for k in HIGH):
            s = 3
        elif any(k in n for k in MID):
            s = 2
        elif any(k in n for k in LOW_NAME):
            s = -1
        if low.endswith((".pdf", ".hwp", ".hwpx", ".doc", ".docx")):
            s += 1
        scored.append((s, i, n, u))
    scored.sort(key=lambda x: (-x[0], x[1]))   # 점수 높은 순, 같으면 원래 순서
    return [(n, u) for _, _, n, u in scored[:limit]]


MAX_ATTACH_BYTES = 30 * 1024 * 1024   # 30MB 초과 파일은 건너뜀


def scan_attachments(item, limit, use_ai=False, notice_name=""):
    """선별된 첨부파일에서 담당자 정보 탐색.
    반환: (phone, ctx, src_file, dept, person, via)  via: 'AI'|'규칙'|''"""
    best = ("", "", "", "", "", "")
    for att_name, att_url in attachments_of(item, limit):
        try:
            r = requests.get(att_url, headers=HEADERS, timeout=60, stream=True)
            if r.status_code != 200:
                continue
            cl = int(r.headers.get("Content-Length") or 0)
            if cl > MAX_ATTACH_BYTES:
                continue
            data = r.content
            if len(data) < 500 or len(data) > MAX_ATTACH_BYTES:
                continue
            text = extract_text(att_name, data)
            if not text:
                continue
            # ① AI 정밀 추출 (키가 있을 때) — 성공 시 즉시 확정
            if use_ai:
                info = ai_extract_contact(text, notice_name, att_name)
                if info:
                    return (info["전화"], info.get("근거", ""), att_name,
                            info.get("부서", ""), info.get("담당자", ""), "AI")
            # ② 규칙 기반 (AI 미사용/실패 시)
            contacts = find_contacts(text)
            if contacts and contacts[0][0] > 0:
                return (contacts[0][1], contacts[0][2], att_name,
                        "", "", "규칙")
            if contacts and not best[0]:
                best = (contacts[0][1], contacts[0][2], att_name,
                        "", "", "규칙")
        except Exception:
            continue
        time.sleep(0.2)
    return best


# ---------- 문서 텍스트 추출 ----------
def text_from_pdf(data):
    try:
        import pdfplumber
        out = []
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for pg in pdf.pages[:30]:
                out.append(pg.extract_text() or "")
        return "\n".join(out)
    except Exception:
        return ""


def text_from_hwpx(data):
    try:
        out = []
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            for name in z.namelist():
                if name.startswith("Contents/section"):
                    xml = z.read(name).decode("utf-8", "ignore")
                    out.append(re.sub(r"<[^>]+>", " ", xml))
        return "\n".join(out)
    except Exception:
        return ""


def text_from_hwp(data):
    try:
        import olefile
        if not olefile.isOleFile(io.BytesIO(data)):
            return ""
        ole = olefile.OleFileIO(io.BytesIO(data))
        header = ole.openstream("FileHeader").read()
        compressed = bool(header[36] & 1)
        chunks = []
        for entry in ole.listdir():
            if entry[0] == "BodyText":
                raw = ole.openstream(entry).read()
                if compressed:
                    try:
                        raw = zlib.decompress(raw, -15)
                    except Exception:
                        continue
                i, n = 0, len(raw)
                while i + 4 <= n:
                    hdr = struct.unpack_from("<I", raw, i)[0]
                    tag = hdr & 0x3FF
                    size = (hdr >> 20) & 0xFFF
                    i += 4
                    if size == 0xFFF:
                        if i + 4 > n:
                            break
                        size = struct.unpack_from("<I", raw, i)[0]
                        i += 4
                    if tag == 67 and i + size <= n:
                        try:
                            chunks.append(raw[i:i + size].decode("utf-16-le", "ignore"))
                        except Exception:
                            pass
                    i += size
        ole.close()
        return re.sub(r"[\x00-\x08\x0b-\x1f]", " ", "\n".join(chunks))
    except Exception:
        return ""


def extract_text(filename, data):
    low = filename.lower()
    if low.endswith(".pdf") or data[:4] == b"%PDF":
        return text_from_pdf(data)
    if low.endswith(".hwpx") or data[:2] == b"PK":
        t = text_from_hwpx(data)
        if t:
            return t
    return text_from_hwp(data)


DEPT_RE = re.compile(r"[가-힣A-Za-z0-9]{2,14}(?:과|팀|사업소|센터|본부|단|국|실|소)\b")
DEPT_STOP = ("결과", "효과", "통과", "초과", "부과", "경과", "성과")


def dept_of(ctx):
    """문맥에서 부서명(○○과, ○○팀 등) 추출 — 전화번호에 가까운 것 우선"""
    cands = [m.group() for m in DEPT_RE.finditer(ctx or "")
             if not m.group().endswith(DEPT_STOP)]
    return cands[-1] if cands else ""


ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"


def get_anthropic_key():
    try:
        return st.secrets["ANTHROPIC_API_KEY"].strip()
    except Exception:
        return None


def trim_for_ai(text, max_chars=6000):
    """전화번호·담당 키워드 주변 발췌로 토큰 절약 (문서 전체 전송 방지)"""
    if len(text) <= max_chars:
        return text
    spots = []
    for kw in ("담당", "문의", "감독", "연락처", "전화"):
        p = 0
        while True:
            p = text.find(kw, p)
            if p < 0 or len(spots) > 20:
                break
            spots.append(p)
            p += len(kw)
    for m in PHONE_RE.finditer(text):
        spots.append(m.start())
        if len(spots) > 40:
            break
    if not spots:
        return text[:max_chars // 2] + "\n...\n" + text[-max_chars // 2:]
    spots.sort()
    chunks, last_end = [], -1
    for s in spots:
        b, e = max(0, s - 400), min(len(text), s + 400)
        if b <= last_end:
            chunks[-1] = (chunks[-1][0], max(chunks[-1][1], e))
        else:
            chunks.append((b, e))
        last_end = chunks[-1][1]
    out = "\n...\n".join(text[b:e] for b, e in chunks)
    return out[:max_chars]


def ai_extract_contact(doc_text, notice_name, filename):
    """Claude(Haiku)로 실무담당 부서·성명·전화 추출.
    반환: {"부서","담당자","전화","구분","근거"} 또는 None(실패/미발견)"""
    key = get_anthropic_key()
    if not key or not doc_text:
        return None
    excerpt = trim_for_ai(doc_text)
    prompt = (
        f"다음은 나라장터 입찰공고 '{notice_name}'의 첨부문서 "
        f"'{filename}'에서 발췌한 내용이다.\n\n"
        "이 공사의 '실무담당자' 정보를 찾아라. 우선순위:\n"
        "1) 공사감독/현장감독/사업담당/실무담당/주무관 등 실제 업무 담당\n"
        "2) 위가 없으면 문의처로 안내된 담당\n"
        "제외: 계약담당·입찰집행·회계 담당은 실무담당이 따로 있으면 제외.\n"
        "팩스번호는 전화로 취급하지 마라.\n\n"
        "반드시 아래 JSON만 출력(설명·백틱 금지). 못 찾으면 전화를 빈값으로:\n"
        '{"부서":"","담당자":"","전화":"","구분":"실무|계약|문의처","근거":"발췌 원문 한 줄"}\n\n'
        f"--- 문서 발췌 ---\n{excerpt}")
    try:
        r = requests.post(
            ANTHROPIC_URL,
            headers={"x-api-key": key,
                     "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": "claude-haiku-4-5", "max_tokens": 300,
                  "messages": [{"role": "user", "content": prompt}]},
            timeout=60)
        if r.status_code != 200:
            return None
        parts = r.json().get("content", [])
        txt = "".join(p.get("text", "") for p in parts
                      if p.get("type") == "text")
        txt = re.sub(r"```(?:json)?|```", "", txt).strip()
        m = re.search(r"\{.*\}", txt, re.S)
        if not m:
            return None
        import json as _json
        info = _json.loads(m.group(0))
        if not str(info.get("전화", "")).strip():
            return None
        info["전화"] = re.sub(r"[\s\.\)]", "-",
                            str(info["전화"])).strip("-")
        info["전화"] = re.sub(r"-{2,}", "-", info["전화"])
        return info
    except Exception:
        return None


def find_contacts(text):
    results = []
    if not text:
        return results
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for idx, line in enumerate(lines):
        for m in PHONE_RE.finditer(line):
            phone = re.sub(r"[\s\.\)]", "-", m.group()).strip("-")
            phone = re.sub(r"-{2,}", "-", phone)
            ctx = " / ".join(lines[max(0, idx - 2): idx + 1])[-200:]
            score = sum(1 for kw in CONTACT_KEYWORDS if kw in ctx)
            if "팩스" in line or "FAX" in line.upper():
                score -= 3
            results.append((score, phone, ctx))
    results.sort(key=lambda x: -x[0])
    seen, out = set(), []
    for s, p, c in results:
        if p not in seen:
            seen.add(p)
            out.append((s, p, c))
    return out[:3]


# ------------------------------------------------------------
# 메일 발송 + 앱 내 자동 실행 스케줄러
# ------------------------------------------------------------
LOG_PATH = "/tmp/g2b_daily_last.log"
MARK_PATH = "/tmp/g2b_daily_mark.txt"


def mail_cfg():
    """Secrets에 메일 설정이 모두 있으면 dict, 아니면 None"""
    try:
        s = st.secrets
        need = ("SMTP_HOST", "SMTP_USER", "SMTP_PASS", "MAIL_TO")
        if not all(k in s for k in need):
            return None
        return {"host": s["SMTP_HOST"],
                "port": int(s.get("SMTP_PORT", 465)),
                "user": s["SMTP_USER"], "pw": s["SMTP_PASS"],
                "to": s["MAIL_TO"],
                "sender": s.get("MAIL_FROM", s["SMTP_USER"])}
    except Exception:
        return None


def send_report_mail(cfg, filename, data, summary_lines):
    msg = EmailMessage()
    today = datetime.date.today().strftime("%Y-%m-%d")
    msg["Subject"] = f"[나라장터] {today} 조회 리포트"
    msg["From"] = cfg["sender"]
    msg["To"] = cfg["to"]
    msg.set_content("\n".join(
        [f"{today} 나라장터 조회 결과입니다.", ""] + list(summary_lines)
        + ["", "- 나라장터 입찰정보 수집기 자동 발송"]))
    if data:
        msg.add_attachment(data, maintype="application",
                           subtype=("vnd.openxmlformats-officedocument"
                                    ".spreadsheetml.sheet"),
                           filename=filename)
    ctx = ssl.create_default_context()
    if cfg["port"] == 465:
        with smtplib.SMTP_SSL(cfg["host"], cfg["port"], context=ctx) as s:
            s.login(cfg["user"], cfg["pw"])
            s.send_message(msg)
    else:
        with smtplib.SMTP(cfg["host"], cfg["port"]) as s:
            s.starttls(context=ctx)
            s.login(cfg["user"], cfg["pw"])
            s.send_message(msg)


def run_daily_job():
    """daily_report.py를 하위 프로세스로 실행 (Secrets → 환경변수 전달)"""
    script = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "daily_report.py")
    if not os.path.exists(script):
        return False, "daily_report.py 파일이 저장소에 없습니다."
    env = dict(os.environ)
    try:
        for k in ("SERVICE_KEY", "ANTHROPIC_API_KEY", "SMTP_HOST",
                  "SMTP_PORT", "SMTP_USER", "SMTP_PASS", "MAIL_TO",
                  "MAIL_FROM", "SEARCH_TYPE", "TASK", "REGION",
                  "PRICE_MIN_UK", "PRICE_MAX_UK", "DAYS_BACK",
                  "INCLUDE_NATIONWIDE", "MAX_ATTACH"):
            if k in st.secrets:
                env[k] = str(st.secrets[k])
    except Exception:
        pass
    env["TZ"] = "Asia/Seoul"
    try:
        r = subprocess.run([sys.executable, script], env=env,
                           capture_output=True, text=True, timeout=1800)
        out = (r.stdout or "") + ("\n" + r.stderr if r.stderr else "")
        stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        with open(LOG_PATH, "w", encoding="utf-8") as f:
            f.write(f"실행시각: {stamp}\n{out[-4000:]}")
        return r.returncode == 0, out[-4000:]
    except Exception as e:
        return False, str(e)


def _kst_now():
    return datetime.datetime.utcnow() + datetime.timedelta(hours=9)


def _scheduler_loop(auto_time):
    hh, mm = (int(x) for x in auto_time.split(":"))
    while True:
        now = _kst_now()
        today_mark = now.strftime("%Y-%m-%d")
        done = ""
        try:
            done = open(MARK_PATH, encoding="utf-8").read().strip()
        except Exception:
            pass
        due = (now.hour, now.minute) >= (hh, mm)
        if due and done != today_mark:
            try:
                open(MARK_PATH, "w", encoding="utf-8").write(today_mark)
                run_daily_job()
            except Exception:
                pass
        time.sleep(60)


@st.cache_resource
def start_scheduler():
    """AUTO_SEND=1이고 메일 설정이 있으면 백그라운드 스케줄러 시작"""
    try:
        if str(st.secrets.get("AUTO_SEND", "0")) != "1" or not mail_cfg():
            return "off"
        auto_time = str(st.secrets.get("AUTO_TIME", "08:00"))
        t = threading.Thread(target=_scheduler_loop, args=(auto_time,),
                             daemon=True)
        t.start()
        return f"on@{auto_time}"
    except Exception as e:
        return f"error:{e}"


def make_excel(headers, rows, sheet):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    for r in rows:
        ws.append(r)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = \
            max(12, min(55, len(str(h)) * 2 + 8))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ------------------------------------------------------------
# 화면 — 검색조건
# ------------------------------------------------------------
st.title("📋 나라장터 입찰정보 수집기")
st.caption("입찰공고·개찰결과 조회 + 공고서 실무담당 연락처 자동 추출")

if get_service_key() is None:
    st.error("인증키가 설정되지 않았습니다. 앱 설정(Settings → Secrets)에 "
             "`SERVICE_KEY = \"발급받은키\"` 를 추가하세요.")
    st.stop()

search_type = st.radio("검색유형", ["입찰공고", "개찰결과"], horizontal=True)

today = datetime.date.today()
date_basis = "개찰일" if search_type == "개찰결과" else "공고게시일"
c1, c2 = st.columns(2)
with c1:
    date_from = st.date_input(f"조회 시작일 ({date_basis} 기준)",
                              today - datetime.timedelta(days=14))
with c2:
    date_to = st.date_input(f"조회 종료일 ({date_basis} 기준)", today)

c3, c4 = st.columns(2)
with c3:
    task = st.selectbox("업무구분", list(TASKS.keys()))
with c4:
    region_name = st.selectbox("참가제한지역", list(REGIONS.keys()),
                               index=list(REGIONS.keys()).index("경상남도"))
    include_nationwide = False
    if region_name not in ("전체", "전국(제한없음)"):
        include_nationwide = st.checkbox(
            "전국(제한없음) 공고도 수요기관이 해당 지역이면 포함 (개찰결과)",
            value=False,
            help="끄면 나라장터 참가제한지역 필터와 동일합니다(빠름). "
                 "켜면 전국 개찰목록을 추가 수신해 수요기관이 선택 지역 소속인 "
                 "전국(제한없음) 공고까지 포함합니다(느려짐).")

c5, c6 = st.columns(2)
with c5:
    price_min_uk = st.number_input("추정가격 하한 (억원)", 0.0, 10000.0, 1.0, step=0.5)
with c6:
    price_max_uk = st.number_input("추정가격 상한 (억원, 0=제한없음)",
                                   0.0, 10000.0, 0.0, step=0.5)

do_contacts = st.checkbox("공고서에서 실무담당 전화번호 추출", value=True)
use_ai = False
if do_contacts:
    if get_anthropic_key():
        use_ai = st.checkbox(
            "🤖 Claude AI 정밀 추출 사용", value=True,
            help="공고서 내용을 Claude(Haiku)가 읽고 실무담당 부서·성명·전화를 "
                 "구분해 추출합니다. 계약담당/팩스 오인이 크게 줄며, "
                 "비용은 공고 1건당 약 5~10원 수준입니다. "
                 "실패 시 자동으로 규칙 기반 추출로 대체됩니다.")
    else:
        st.caption("💡 앱 설정(Secrets)에 ANTHROPIC_API_KEY를 추가하면 "
                   "Claude AI 정밀 추출을 사용할 수 있습니다 "
                   "(부서·성명까지 구분, 정확도 향상)")
max_attach = st.slider("공고당 첨부파일 분석 개수", 1, 5, 3,
                       help="많을수록 정확하지만 느려집니다") if do_contacts else 0
if search_type == "개찰결과":
    st.caption("※ 개찰결과는 공고번호로 원 입찰공고를 찾아 담당자 연락처를 함께 추출합니다. "
               "'나라장터 낙찰정보서비스' 활용신청이 된 인증키여야 하며, "
               "건마다 공고 조회가 추가되어 시간이 더 걸립니다.")

# ------------------------------------------------------------
# 자동 발송 설정 (앱 내)
# ------------------------------------------------------------
sched_state = start_scheduler()
with st.expander("⏰ 매일 자동 조회·메일 발송 설정", expanded=False):
    mc = mail_cfg()
    if mc:
        st.success(f"메일 설정 완료 → 수신: {mc['to']}")
    else:
        st.info("앱 설정(Settings → Secrets)에 아래 항목을 추가하면 "
                "메일 발송이 활성화됩니다:")
        st.code('SMTP_HOST = "smtp.gmail.com"   # 네이버는 smtp.naver.com\n'
                'SMTP_PORT = "465"\n'
                'SMTP_USER = "보내는메일@gmail.com"\n'
                'SMTP_PASS = "앱 비밀번호"      # 계정 비밀번호 아님!\n'
                'MAIL_TO   = "받는메일@company.com"', language="toml")
    if sched_state.startswith("on@"):
        st.success(f"자동 실행 켜짐 — 매일 {sched_state[3:]} (한국시간), "
                   f"어제~오늘 개찰분을 조회해 발송합니다")
    else:
        st.info("자동 실행을 켜려면 Secrets에 추가: "
                '`AUTO_SEND = "1"`, `AUTO_TIME = "08:00"`')
    st.caption("⚠️ 무료 Streamlit 호스팅은 접속이 없으면 앱이 잠들어 자동 실행이 "
               "건너뛰어질 수 있습니다. uptimerobot.com(무료)에서 이 앱 주소를 "
               "5분 간격 모니터로 등록해 두면 항상 깨어 있습니다.")
    if os.path.exists(LOG_PATH):
        with open(LOG_PATH, encoding="utf-8") as f:
            st.text_area("최근 자동 실행 로그", f.read(), height=150)
    if mc and st.button("▶ 지금 바로 조회+메일 발송 (테스트)"):
        with st.spinner("백그라운드 조회·발송 실행 중... (수 분 소요)"):
            ok2, out = run_daily_job()
        (st.success if ok2 else st.error)(
            ("발송 완료! 메일함을 확인하세요." if ok2 else "실패 — 아래 로그 확인"))
        st.text_area("실행 로그", out, height=200)

# 직전 조회 결과 메일 발송 (조회 후 세션에 보관된 파일)
_mc = mail_cfg()
if _mc and st.session_state.get("last_report"):
    _fn, _sum = (st.session_state["last_report"][0],
                 st.session_state["last_report"][2])
    if st.button(f"📧 방금 조회 결과 메일로 발송 ({_fn})"):
        try:
            send_report_mail(_mc, _fn, st.session_state["last_report"][1],
                             _sum)
            st.success(f"발송 완료 → {_mc['to']}")
        except Exception as e:
            st.error(f"메일 발송 실패: {e}")

# ------------------------------------------------------------
# 실행
# ------------------------------------------------------------
if st.button("🔍 조회 시작", type="primary", use_container_width=True):
    if date_from > date_to:
        st.error("조회 시작일이 종료일보다 늦습니다.")
        st.stop()

    begin = date_from.strftime("%Y%m%d") + "0000"
    end = date_to.strftime("%Y%m%d") + "2359"
    p_min = int(price_min_uk * 100_000_000)
    p_max = int(price_max_uk * 100_000_000)  # 0이면 제한없음
    region_codes = REGIONS[region_name]

    extra = {"inqryDiv": 1, "inqryBgnDt": begin, "inqryEndDt": end}
    log = st.empty()

    if search_type == "입찰공고":
        op = TASKS[task][0]
        with st.spinner("입찰공고 목록 조회 중..."):
            if region_codes:   # 지역코드별로 조회 후 병합 (통합·개편 구코드 포함)
                items, err, seen = [], None, set()
                for code in region_codes:
                    part, e = call_api(BID_BASES, op,
                                       {**extra, "prtcptLmtRgnCd": code}, log)
                    err = err or e
                    for it in part:
                        k = (it.get("bidNtceNo"), it.get("bidNtceOrd"))
                        if k not in seen:
                            seen.add(k)
                            items.append(it)
            else:              # 전체 / 전국(제한없음): 코드 필터 없이 조회
                items, err = call_api(BID_BASES, op, extra, log)
        if items and region_name == "전국(제한없음)":
            # 참가제한이 걸리지 않은(전국) 공고만 남김
            items = [it for it in items
                     if region_match(region_name,
                                     fetch_psbl_rgn(it.get("bidNtceNo")), it)]
    else:
        # 지역이 선택되면 전국 개찰목록을 받을 필요가 없다(공고 기반 역방향 수집).
        # 전국 목록은 '전체/전국(제한없음)' 선택 또는 전국공고 보충 옵션일 때만 수신.
        need_open_list = (region_name in ("전체", "전국(제한없음)")
                          or include_nationwide)
        items, err = [], None
        if need_open_list:
            op = TASKS[task][1]
            # 낙찰정보 API의 기간 조회는 '등록일시' 기준이라 나라장터 화면의
            # '개찰일자' 기준과 어긋난다 → 넉넉한 기간으로 받은 뒤 개찰일로 필터.
            wide = {"inqryDiv": 1,
                    "inqryBgnDt": (date_from - datetime.timedelta(days=3)
                                   ).strftime("%Y%m%d") + "0000",
                    "inqryEndDt": (date_to + datetime.timedelta(days=2)
                                   ).strftime("%Y%m%d") + "2359"}
            with st.spinner("개찰결과 목록 조회 중..."):
                items, err = call_api(SCSBID_BASES, op, wide, log)
            if items:
                n_all = len(items)
                items = [it for it in items
                         if (openg_date_of(it) is None)    # 개찰일 판독불가 건은 유지
                         or (date_from <= openg_date_of(it) <= date_to)]
                if n_all != len(items):
                    st.caption(f"개찰일 {date_from}~{date_to} 범위 밖 "
                               f"{n_all - len(items)}건 제외")

    log.empty()
    if err and not items and (search_type == "입찰공고" or need_open_list):
        st.error(f"API 조회 실패: {err}")
        if search_type == "개찰결과":
            st.info("공공데이터포털에서 '조달청_나라장터 낙찰정보서비스' "
                    "활용신청이 승인되었는지 확인하세요.")
        else:
            st.info("인증키 활성화 전이거나(발급 후 몇 시간 소요), "
                    "'나라장터 입찰공고정보서비스' 활용신청을 확인하세요.")
        st.stop()

    # 금액 필터
    # ※ 개찰결과 목록 API에는 추정가격 필드가 없으므로(모두 0으로 계산됨)
    #   목록 단계에서는 필터하지 않고, 아래 루프에서 원 공고의 추정가격으로 필터한다.
    if search_type == "입찰공고":
        # 공사 등은 목록 응답에 추정가격이 없음 → 기초금액 오퍼레이션으로 보충
        bsis_map = {}
        if TASKS_BSIS.get(task) and any(price_of(it) == 0 for it in items):
            with st.spinner("기초금액(추정가격) 조회 중..."):
                brows, _ = call_api_chunked(
                    BID_BASES, TASKS_BSIS[task], {"inqryDiv": 1},
                    date_from, date_to, log)
            log.empty()
            for b in brows:
                p = price_of(b)
                if p:
                    bsis_map[str(b.get("bidNtceNo", ""))] = p

        def eff_price(it):
            return price_of(it) or bsis_map.get(str(it.get("bidNtceNo", "")), 0)

        filtered = [it for it in items
                    if (eff_price(it) == 0)      # 가격 미확인 건은 포함
                    or (eff_price(it) >= p_min
                        and (p_max == 0 or eff_price(it) <= p_max))]
        st.success(f"조건 충족 {len(filtered)}건 (전체 수신 {len(items)}건)")
        n_unk = sum(1 for it in filtered if eff_price(it) == 0)
        if n_unk:
            st.caption(f"추정가격 미확인 {n_unk}건 포함")
    else:
        filtered = items
        if need_open_list:
            st.success(f"개찰일 기준 수신 {len(items)}건 — 건별로 참가제한지역·"
                       f"추정가격을 확인해 필터를 적용합니다")
    if not filtered and not (search_type == "개찰결과"
                             and not need_open_list):
        st.stop()

    fname_date = today.strftime("%Y%m%d")

    # ---------------- 입찰공고 결과 ----------------
    if search_type == "입찰공고":
        rows = []
        progress = st.progress(0.0)
        status = st.empty()
        for i, item in enumerate(filtered, 1):
            name = item.get("bidNtceNm", "")
            best_phone, best_ctx, src_file = "", "", ""
            ai_dept, ai_person, via = "", "", ""
            if do_contacts:
                status.write(f"[{i}/{len(filtered)}] {name[:35]}... 공고서 분석 중")
                best_phone, best_ctx, src_file, ai_dept, ai_person, via = \
                    scan_attachments(item, max_attach, use_ai, name)
            dept = ai_dept or dept_of(best_ctx)
            if ai_person:
                dept = f"{dept} {ai_person}".strip()
            rows.append([item.get("bidNtceNo", ""), item.get("bidNtceOrd", ""),
                         name, item.get("dminsttNm", ""), eff_price(item),
                         dept, best_phone, best_ctx, src_file,
                         item.get("ntceInsttOfclNm", ""),
                         item.get("ntceInsttOfclTelNo", ""),
                         item.get("bidNtceDtlUrl") or item.get("bidNtceUrl") or "",
                         ("AI추출" if best_phone and via == "AI" else
                          "추출성공" if best_phone else
                          ("수동확인필요" if do_contacts else "-"))])
            progress.progress(i / len(filtered))
        status.empty()

        if do_contacts:
            ok = sum(1 for r in rows if r[-1] in ("추출성공", "AI추출"))
            st.success(f"완료! 자동 추출 {ok}건 / 수동확인 필요 {len(rows) - ok}건")

        st.dataframe(
            [{"공고명": r[2][:30], "수요기관": r[3],
              "실무담당 부서": r[5], "실무담당 전화": r[6],
              "집행관(계약) 전화": r[10],
              "추정가격(억)": round(r[4] / 1e8, 2),
              "공고 바로가기": r[11], "상태": r[-1]} for r in rows],
            use_container_width=True, hide_index=True,
            column_config={"공고 바로가기": st.column_config.LinkColumn(
                "공고 바로가기", display_text="열기")})

        headers = ["공고번호", "차수", "공고명", "수요기관", "추정가격(원)",
                   "실무담당 부서", "실무담당 전화", "추출 문맥", "출처 파일",
                   "집행관(계약담당)", "집행관 전화", "공고 상세URL", "상태"]
        _xls = make_excel(headers, rows, "입찰공고")
        _fn = f"입찰공고_{task}_{region_name}_{fname_date}.xlsx"
        st.session_state["last_report"] = (
            _fn, _xls.getvalue() if hasattr(_xls, "getvalue") else _xls,
            [f"입찰공고 {len(rows)}건 (지역: {region_name})"])
        st.download_button("📥 엑셀 다운로드", data=_xls, file_name=_fn,
                           mime="application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet",
                           use_container_width=True)

    # ---------------- 개찰결과 결과 ----------------
    else:
        # ⚡ 역방향 수집: 나라장터 웹과 동일하게, 지역으로 서버 필터된
        # '입찰공고'에서 개찰일·금액 조건에 맞는 후보를 먼저 확정한 뒤
        # 그 후보들에 대해서만 개찰결과 조회와 공고서 분석을 수행한다.
        # (개찰결과 API는 지역·금액 검색을 지원하지 않으므로 전국 목록을
        #  받는 대신 이 방식이 정확하고 빠르다)
        stats = {"개찰일 범위 밖": 0, "지역 불일치": 0, "금액 범위 밖": 0,
                 "지역정보 미확인(기관명 판정)": 0, "추정가격 미확인(포함)": 0}
        cands = []       # (notice or None, open_item or None) 후보 목록
        seen_no = set()

        if region_name not in ("전체", "전국(제한없음)") and region_codes:
            # ① 지역 제한 공고를 '개찰일시 기준'으로 직접 검색 (나라장터
            #    화면과 동일한 조회구분) → 해당 기간 것만 수신되어 빠름
            notices, nerr = [], None
            with st.spinner(f"{region_name} 제한 공고 조회 중(개찰일 기준)..."):
                for code in region_codes:
                    part, e = call_api(
                        BID_BASES, TASKS[task][0],
                        {"inqryDiv": 2,          # 2 = 개찰일시 기준
                         "inqryBgnDt": begin, "inqryEndDt": end,
                         "prtcptLmtRgnCd": code}, log)
                    nerr = nerr or e
                    for n in part:
                        nno = str(n.get("bidNtceNo", ""))
                        if nno and nno not in seen_no:
                            seen_no.add(nno)
                            notices.append(n)
            if not notices:
                # 폴백: 개찰일시 조회 미지원/무응답 시 게시일 기준 120일 수집
                st.caption("개찰일시 기준 조회 결과가 없어 게시일 기준으로 "
                           "재수집합니다(시간이 더 걸립니다)")
                with st.spinner(f"{region_name} 제한 공고 조회 중(게시일 기준)..."):
                    for code in region_codes:
                        part, e = call_api_chunked(
                            BID_BASES, TASKS[task][0],
                            {"inqryDiv": 1, "prtcptLmtRgnCd": code},
                            date_from - datetime.timedelta(days=120),
                            date_to, log)
                        nerr = nerr or e
                        for n in part:
                            nno = str(n.get("bidNtceNo", ""))
                            if nno and nno not in seen_no:
                                seen_no.add(nno)
                                notices.append(n)
            log.empty()
            if not notices and nerr:
                st.error(f"{region_name} 제한 공고 조회 실패: {str(nerr)[:200]}")
                st.stop()
            # ② 개찰일 이중 확인 (판독 불가 건은 서버 필터를 신뢰하고 포함)
            for n in notices:
                d = openg_date_of(n)
                if d is None or (date_from <= d <= date_to):
                    cands.append((n, None))
                else:
                    stats["개찰일 범위 밖"] += 1
            st.success(f"{region_name} 제한 공고 {len(notices)}건 중 "
                       f"개찰일 {date_from}~{date_to} 해당 {len(cands)}건")
            # ③ (옵션) 전국(제한없음) 공고 보충: 전국 개찰목록에서
            #    수요기관이 지역 소속인 건만 골라 정밀 확인
            if include_nationwide and filtered:
                kws = INSTT_KEYWORDS.get(region_name, [region_name])
                extra_c = 0
                for it in filtered:
                    no = str(it.get("bidNtceNo", ""))
                    if no in seen_no:
                        continue
                    blob = (str(it.get("dminsttNm", "")) +
                            str(it.get("ntceInsttNm", "")))
                    if not any(k in blob for k in kws):
                        continue
                    rgn = fetch_psbl_rgn(no)
                    is_nw = (rgn == []) or (rgn and any(
                        k in " ".join(rgn)
                        for k in RGN_NAME_KEYWORDS["전국(제한없음)"]))
                    if rgn is None:
                        stats["지역정보 미확인(기관명 판정)"] += 1
                        is_nw = True     # 확인 불가 + 기관명 일치 → 포함
                    if is_nw:
                        seen_no.add(no)
                        cands.append((None, it))
                        extra_c += 1
                if extra_c:
                    st.caption(f"전국(제한없음) 공고 보충 {extra_c}건 추가")
        else:
            # '전체' 또는 '전국(제한없음)' 선택: 전국 개찰목록 기반 (기존 방식)
            for it in filtered:
                no = str(it.get("bidNtceNo", ""))
                if region_name == "전국(제한없음)":
                    rgn = fetch_psbl_rgn(no)
                    if rgn is None:
                        stats["지역정보 미확인(기관명 판정)"] += 1
                    if not region_match(region_name, rgn, it):
                        stats["지역 불일치"] += 1
                        continue
                cands.append((None, it))

        if not cands:
            st.warning("조건에 해당하는 개찰 건이 없습니다.")
            st.caption(" · ".join(f"{k} {v}건" for k, v in stats.items() if v))
            st.stop()

        # ④ 후보별 처리: 금액 확인 → 개찰결과 조회 → 공고서 분석
        rows = []
        progress = st.progress(0.0)
        status = st.empty()
        for i, (notice, op_it) in enumerate(cands, 1):
            src = notice if notice is not None else op_it
            no = str(src.get("bidNtceNo", ""))
            name = (src.get("bidNtceNm", "")
                    or src.get("prdctClsfcNoNm", ""))
            best_phone, best_ctx, src_file = "", "", ""
            ai_dept, ai_person, via = "", "", ""
            ofcl_nm, ofcl_tel, url = "", "", ""
            # 원 공고 확보 (보충 경로는 여기서 조회)
            if notice is None:
                status.write(f"[{i}/{len(cands)}] {name[:35]}... 원 공고 조회 중")
                notice = fetch_notice(no, task)
            # 금액 필터 (공고 → 기초금액 순, 미확인 건은 포함)
            prc = price_of(notice) if notice else 0
            if not prc:
                status.write(f"[{i}/{len(cands)}] {name[:35]}... 기초금액 조회 중")
                prc = fetch_bsis_price(no, task)
            if prc and (prc < p_min or (p_max and prc > p_max)):
                stats["금액 범위 밖"] += 1
                progress.progress(i / len(cands))
                continue
            if not prc:
                stats["추정가격 미확인(포함)"] += 1
            # 개찰결과 확보 (역방향 경로는 여기서 공고번호로 조회)
            if op_it is None:
                status.write(f"[{i}/{len(cands)}] {name[:35]}... 개찰결과 조회 중")
                op_it = fetch_openg(no, task) or {}
            # 담당자 정보 + 공고서 분석 (조건 확정 건만)
            if notice:
                ofcl_nm = notice.get("ntceInsttOfclNm", "")
                ofcl_tel = notice.get("ntceInsttOfclTelNo", "")
                url = (notice.get("bidNtceDtlUrl")
                       or notice.get("bidNtceUrl") or "")
                if do_contacts:
                    status.write(f"[{i}/{len(cands)}] {name[:35]}... 공고서 분석 중")
                    (best_phone, best_ctx, src_file,
                     ai_dept, ai_person, via) = \
                        scan_attachments(notice, max_attach, use_ai, name)
            state = ("AI추출" if best_phone and via == "AI" else
                     "추출성공" if best_phone else
                     ("공고미확인" if not notice else
                      ("수동확인필요" if do_contacts else "-")))
            rows.append([no, name,
                         src.get("dminsttNm", "") or src.get("ntceInsttNm", ""),
                         src.get("opengDt", "") or src.get("rlOpengDt", ""),
                         op_it.get("bidwinnrNm", "")
                         or op_it.get("opengCorpInfo", ""),
                         prc, op_it.get("sucsfbidRate", ""),
                         ((f"{ai_dept} {ai_person}".strip())
                          if (ai_dept or ai_person) else dept_of(best_ctx)),
                         best_phone, best_ctx, src_file,
                         ofcl_nm, ofcl_tel, url, state])
            progress.progress(i / len(cands))
            time.sleep(0.2)
        status.empty()

        st.success(f"조건 충족 {len(rows)}건 (후보 {len(cands)}건 중)")
        st.caption(" · ".join(f"{k} {v}건" for k, v in stats.items() if v))
        if not rows:
            st.stop()
        if do_contacts:
            ok = sum(1 for r in rows if r[-1] in ("추출성공", "AI추출"))
            st.success(f"자동 추출 {ok}건 / 수동확인 필요 {len(rows) - ok}건")

        st.dataframe(
            [{"공고명": r[1][:30], "수요기관": r[2], "개찰일": str(r[3])[:10],
              "낙찰업체": r[4], "추정가격(억)": round(r[5] / 1e8, 2),
              "실무담당 부서": r[7], "실무담당 전화": r[8],
              "집행관(계약) 전화": r[12],
              "공고 바로가기": r[13], "상태": r[-1]} for r in rows],
            use_container_width=True, hide_index=True,
            column_config={"공고 바로가기": st.column_config.LinkColumn(
                "공고 바로가기", display_text="열기")})

        headers = ["공고번호", "공고명", "수요기관", "개찰일시",
                   "낙찰업체", "추정가격(원)", "낙찰률(%)",
                   "실무담당 부서", "실무담당 전화", "추출 문맥", "출처 파일",
                   "집행관(계약담당)", "집행관 전화", "공고 상세URL", "상태"]
        _xls = make_excel(headers, rows, "개찰결과")
        _fn = f"개찰결과_{task}_{region_name}_{fname_date}.xlsx"
        st.session_state["last_report"] = (
            _fn, _xls.getvalue() if hasattr(_xls, "getvalue") else _xls,
            [f"개찰결과 {len(rows)}건 (지역: {region_name}, "
             f"개찰일 {date_from}~{date_to})"])
        st.download_button("📥 엑셀 다운로드", data=_xls, file_name=_fn,
                           mime="application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet",
                           use_container_width=True)
